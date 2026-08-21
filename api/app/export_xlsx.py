"""Excel export of the Matching Customers preview — one row per event, the
payload fields flattened into columns (the on-screen table shows the payload
only behind a per-row View toggle; the export is where you get it in bulk).

Works for both windows: days=None exports the live next-run selection,
days=N the trailing-N-days history. Strictly display-side — reads the same
view/TVF the preview reads and can never send a webhook.
"""

import datetime as dt
import io
import json
import re
from concurrent.futures import ThreadPoolExecutor

from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from . import affected

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_PACIFIC = ZoneInfo("America/Los_Angeles")

# Far above any real preview (largest known: tb_signup 90d ≈ 10.5k). If a
# window somehow exceeds it, the file says so in its last row rather than
# silently truncating.
EXPORT_MAX = 20_000

# Identity columns lead, in the order the on-screen table implies; remaining
# payload keys follow in first-seen payload order.
_LEAD = ["email", "first_name", "last_name"]


def _pacific(iso: str | None):
    """ISO timestamp -> naive Pacific datetime (Excel-sortable, matches the
    UI's Pacific rendering). Naive because Excel has no timezone concept."""
    if not iso:
        return None
    try:
        t = dt.datetime.fromisoformat(iso)
    except ValueError:
        return iso
    if t.tzinfo is None:
        t = t.replace(tzinfo=dt.timezone.utc)
    return t.astimezone(_PACIFIC).replace(tzinfo=None)


def _cell(key: str, v):
    """Payload value -> Excel cell value."""
    if v is None:
        return None
    # The one epoch field in the membership payload — raw seconds are useless
    # to a marketer, so it lands as a Pacific datetime like event_at.
    if key == "ticketing_event_date" and isinstance(v, (int, float)) and not isinstance(v, bool):
        return dt.datetime.fromtimestamp(v, tz=dt.timezone.utc).astimezone(_PACIFIC).replace(tzinfo=None)
    if isinstance(v, str):
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
            return dt.date.fromisoformat(v)  # close_date etc. sort as dates
        if v[:1] in "=+-@":
            # openpyxl writes a leading-= string as a live formula; the
            # apostrophe is Excel's own text-literal marker.
            return "'" + v
    return v


def _filename(base: str, part: str) -> str:
    """`<base>--<part>--<YYYY-MM-DD-HHMM>.xlsx`, Pacific-stamped. base is the
    campaign slug when one exists — the same name the portal URL leads with —
    with the trigger key as fallback."""
    stamp = dt.datetime.now(_PACIFIC).strftime("%Y-%m-%d-%H%M")
    safe = re.sub(r'[^A-Za-z0-9._-]', "-", base)
    return f"{safe}--{part}--{stamp}.xlsx"


def preview_xlsx(
    trigger_key: str,
    days: int | None = None,
    q: str | None = None,
    filename_base: str | None = None,
) -> tuple[str, bytes]:
    """(filename, xlsx bytes) for a trigger's preview window."""
    rows, total = affected._preview_rows(trigger_key, q, EXPORT_MAX, 0, days)

    wb = Workbook()
    ws = wb.active
    ws.title = trigger_key[:31]  # Excel's sheet-name limit
    _fill_sheet(ws, rows, total)

    buf = io.BytesIO()
    wb.save(buf)
    window = f"last-{days}-days" if days else "next-run"
    return _filename(filename_base or trigger_key, window), buf.getvalue()


def campaign_xlsx(
    trigger_key: str,
    history_days: int = 90,
    q: str | None = None,
    filename_base: str | None = None,
) -> tuple[str, bytes]:
    """(filename, xlsx bytes) with BOTH windows as worksheet tabs — "Next Run"
    and "Last N Days" — so one file carries the whole preview. A trigger
    without a history branch still exports; its history tab says why it's
    empty instead of pretending zero events."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Next Run"
    rows, total = affected._preview_rows(trigger_key, q, EXPORT_MAX, 0, None)
    _fill_sheet(ws, rows, total)

    ws_hist = wb.create_sheet(f"Last {history_days} Days")
    if trigger_key in affected.HISTORY_TRIGGERS:
        rows, total = affected._preview_rows(trigger_key, q, EXPORT_MAX, 0, history_days)
        _fill_sheet(ws_hist, rows, total)
    else:
        ws_hist.append([
            affected.NO_HISTORY_REASON.get(trigger_key)
            or "No history view for this trigger — only the live next-run selection exists."
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return _filename(filename_base or trigger_key, "campaign"), buf.getvalue()


def all_campaigns_xlsx(
    directory: list[dict], history_days: int = 90
) -> tuple[str, bytes]:
    """(filename, xlsx bytes) with ONE worksheet per campaign, both windows
    stacked in it and told apart by a leading `window` column — "future"
    (the live next-run selection) vs "past_<N>days" (the trailing history).
    directory rows are affected.trigger_directory() entries; sheets are named
    by trigger label (the short human name — slugs blow Excel's 31-char
    sheet-name limit), falling back to slug/key, deduped when truncation
    collides."""
    # Every (trigger, window) fetch in flight at once — serial execution took
    # ~35s across 11 warehouse queries; the BQ client is thread-safe.
    with ThreadPoolExecutor(max_workers=8) as pool:
        fetched = {
            (t["key"], days): pool.submit(affected._preview_rows, t["key"], None, EXPORT_MAX, 0, days)
            for t in directory
            for days in ([None, history_days] if t["key"] in affected.HISTORY_TRIGGERS else [None])
        }
        results = {k: f.result() for k, f in fetched.items()}

    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for t in directory:
        ws = wb.create_sheet(_sheet_name(t, used))
        rows, total = results[(t["key"], None)]
        windows = ["future"] * len(rows)
        if t["key"] in affected.HISTORY_TRIGGERS:
            hist_rows, hist_total = results[(t["key"], history_days)]
            rows = rows + hist_rows
            windows += [f"past_{history_days}days"] * len(hist_rows)
            total += hist_total
        _fill_sheet(ws, rows, total, windows=windows)
        if t["key"] not in affected.HISTORY_TRIGGERS:
            ws.append([])
            ws.append([
                "NOTE: future window only — "
                + (
                    affected.NO_HISTORY_REASON.get(t["key"])
                    or "no history view exists for this trigger."
                )
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return _filename("all-campaigns", "windows"), buf.getvalue()


def _sheet_name(t: dict, used: set[str]) -> str:
    """Excel-safe worksheet name: label > slug > key, illegal chars swapped,
    31-char limit, collisions numbered."""
    base = re.sub(r"[\[\]:*?/\\']", "-", t.get("label") or t.get("slug") or t["key"])[:31]
    name, n = base, 2
    while name.lower() in used:
        name = f"{base[:28]}~{n}"
        n += 1
    used.add(name.lower())
    return name


def _fill_sheet(ws, rows: list[dict], total: int, windows: list[str] | None = None) -> None:
    """Write one preview window into a worksheet: header, data rows, cap
    note, number formats, fitted column widths. `windows` (parallel to rows)
    prepends a window column for sheets that stack both windows."""
    # Column union across every row's payload, first-seen order, identity first.
    cols = list(_LEAD)
    payloads: list[dict] = []
    for r in rows:
        p = {}
        if r.get("payload_json"):
            try:
                p = json.loads(r["payload_json"])
            except ValueError:
                p = {}
        payloads.append(p)
        for k in p:
            # dedup_key rides in the payload too — it gets the dedicated
            # last column instead of a duplicate here.
            if k not in cols and k != "dedup_key":
                cols.append(k)

    header = (["window"] if windows is not None else []) + ["event_at (PT)"] + cols + ["dedup_key"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"

    # Widths tracked while writing (first 500 rows — cosmetic, not worth 20k).
    # NEVER re-scan with iter_rows(max_row=N): in write mode it MATERIALIZES
    # every cell it touches, padding the sheet with hundreds of blank rows.
    widths = [len(str(h)) for h in header]
    for i, (r, p) in enumerate(zip(rows, payloads)):
        # Identity falls back to the row's own columns when a payload lacks it.
        merged = {**{k: r.get(k) for k in _LEAD}, **p}
        values = (
            ([windows[i]] if windows is not None else [])
            + [_pacific(r.get("event_at"))]
            + [_cell(k, merged.get(k)) for k in cols]
            + [r.get("dedup_key")]
        )
        ws.append(values)
        if i < 500:
            for j, v in enumerate(values):
                if v is not None:
                    widths[j] = max(widths[j], len(str(v)))
    if total > len(rows):
        ws.append([])
        ws.append([f"NOTE: {total - len(rows):,} more rows beyond the {EXPORT_MAX:,}-row export cap"])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, dt.datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif isinstance(cell.value, dt.date):
                cell.number_format = "yyyy-mm-dd"

    for j, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = min(42, width + 2)
